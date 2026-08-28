"""
generate_veolia_quotes.py
--------------------------
Generates 4 realistic supplier quotation files (PDF + Excel) for the
Veolia Wastewater Treatment Facility demo — RFQ-WWT-2026-0847.

Supplier data matches the manager's demo spec exactly:
  Supplier A (Gulf Process Systems)      — $4,580/unit — 34 days — Net 30 — 12m warranty
  Supplier B (AquaFlow Controls)         — $4,720/unit — 19 days — Net 45 — 24m warranty  ← AI recommends
  Supplier C (Houston Pump Solutions)    — $4,950/unit — 22 days — Net 45 — 24m warranty
  Supplier D (FlowTech USA)              — $4,490/unit — 28 days — Net 30 — 18m warranty  ← New / Oppora

Run:
  cd backend
  python generate_veolia_quotes.py
"""

import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Output directory ─────────────────────────────────────────────────────────
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'veolia_demo_quotes')
os.makedirs(OUT_DIR, exist_ok=True)

# ── Supplier definitions ──────────────────────────────────────────────────────
SUPPLIERS = [
    {
        "name":          "Gulf Process Systems",
        "label":         "Supplier A",
        "address":       "4200 Westheimer Rd, Suite 310, Houston, TX 77027, USA",
        "email":         "sales@gulfprocesssystems.com",
        "phone":         "+1 (713) 555-0141",
        "country":       "USA",
        "unit_price":    4580.00,
        "qty":           12,
        "currency":      "USD",
        "lead_time":     34,
        "payment_terms": "Net 30",
        "incoterms":     "DAP Houston, TX",
        "warranty":      "12 months",
        "validity":      "60 days",
        "model":         "GP-D120-7B-PVDF",
        "manufacturer":  "Gulf Process Systems Inc.",
        "source":        "Existing Supplier",
        "note":          "Compliant with NEMA 4X. Lead time exceeds 21-day project requirement.",
        "color_rgb":     (0.09, 0.17, 0.27),
    },
    {
        "name":          "AquaFlow Controls",
        "label":         "Supplier B",
        "address":       "8900 Brittmoore Rd, Houston, TX 77041, USA",
        "email":         "procurement@aquaflowcontrols.com",
        "phone":         "+1 (832) 555-0289",
        "country":       "USA",
        "unit_price":    4720.00,
        "qty":           12,
        "currency":      "USD",
        "lead_time":     19,
        "payment_terms": "Net 45",
        "incoterms":     "DDP Houston, TX",
        "warranty":      "24 months",
        "validity":      "60 days",
        "model":         "AF-MD120-PVDF-460V",
        "manufacturer":  "AquaFlow Controls LLC",
        "source":        "Preferred Supplier",
        "note":          "Meets 21-day delivery requirement. Full PVDF/PTFE wetted parts. 4-20mA + local control.",
        "color_rgb":     (0.04, 0.45, 0.18),
    },
    {
        "name":          "MetroChem Systems",
        "label":         "Supplier C",
        "address":       "12330 Cutten Rd, Houston, TX 77066, USA",
        "email":         "sales@metrochemsystems.com",
        "phone":         "+1 (281) 555-0374",
        "country":       "USA",
        "unit_price":    4950.00,
        "qty":           12,
        "currency":      "USD",
        "lead_time":     22,
        "payment_terms": "Net 45",
        "incoterms":     "DAP Houston, TX",
        "warranty":      "24 months",
        "validity":      "60 days",
        "model":         "HPS-CM120-PTFE-7B",
        "manufacturer":  "MetroChem Systems Corp.",
        "source":        "Existing Supplier",
        "note":          "1-day over 21-day requirement. PTFE wetted materials. Includes O&M manual and drawings.",
        "color_rgb":     (0.09, 0.17, 0.55),
    },
    {
        "name":          "Precision Dosing Systems",
        "label":         "Supplier D (New — Oppora Sourced)",
        "address":       "2150 Citywest Blvd, San Antonio, TX 78230, USA",
        "email":         "quotes@precisiondosing.com",
        "phone":         "+1 (210) 555-0412",
        "country":       "USA",
        "unit_price":    4490.00,
        "qty":           12,
        "currency":      "USD",
        "lead_time":     28,
        "payment_terms": "Net 30",
        "incoterms":     "FOB San Antonio, TX",
        "warranty":      "18 months",
        "validity":      "45 days",
        "model":         "FT-DOSE120-460-PVDF",
        "manufacturer":  "Precision Dosing Systems Ltd.",
        "source":        "New Supplier — Oppora Market Discovery",
        "note":          "Lowest unit price. Lead time unverified — stated 28 days. Compliance review required before PO.",
        "color_rgb":     (0.55, 0.15, 0.05),
    },
]

RFQ_REF    = "RFQ-WWT-2026-0847"
ITEM_NAME  = "Chemical Dosing Pump Assembly — Sodium Hypochlorite / Water Treatment"
ITEM_SPEC  = "Motor-driven metering pump | 0–120 L/hr | ≥7 bar | PVDF/PTFE | 460V/3Ph/60Hz | NEMA 4X | 4–20mA"
BUYER      = "Veolia Water Technologies"
DEPT       = "Operations / Procurement"
DELIVERY   = "Houston, Texas, USA"
QUOTE_DATE = "2026-08-28"


# ─────────────────────────────────────────────────────────────────────────────
# PDF GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf(s, file_path):
    c = canvas.Canvas(file_path, pagesize=letter)
    W, H = letter  # 612 x 792

    r, g, b = s["color_rgb"]

    # ── Header Banner ──
    c.setFillColorRGB(r, g, b)
    c.rect(0, H - 80, W, 80, fill=1, stroke=0)

    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(36, H - 38, s["name"].upper())
    c.setFont("Helvetica", 9)
    c.drawString(36, H - 54, "OFFICIAL SUPPLIER QUOTATION — CHEMICAL DOSING PUMPS")
    c.drawString(36, H - 68, f"Source: {s['source']}")

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(W - 36, H - 38, f"Quote Ref: QT-2026-{s['name'][:3].upper()}-847")
    c.setFont("Helvetica", 8.5)
    c.drawRightString(W - 36, H - 52, f"Date: {QUOTE_DATE}")
    c.drawRightString(W - 36, H - 65, f"Valid: {s['validity']} from issue date")

    # ── Vendor / Buyer Info ──
    y = H - 105
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 9.5)
    c.drawString(36, y, "VENDOR:")
    c.setFont("Helvetica", 8.5)
    c.drawString(36, y - 14, s["name"])
    c.drawString(36, y - 26, s["address"])
    c.drawString(36, y - 38, f"Email: {s['email']}")
    c.drawString(36, y - 50, f"Phone: {s['phone']}")

    c.setFont("Helvetica-Bold", 9.5)
    c.drawString(340, y, "PREPARED FOR:")
    c.setFont("Helvetica", 8.5)
    c.drawString(340, y - 14, BUYER)
    c.drawString(340, y - 26, f"Department: {DEPT}")
    c.drawString(340, y - 38, f"Delivery Location: {DELIVERY}")
    c.drawString(340, y - 50, f"RFQ Reference: {RFQ_REF}")

    # Divider
    c.setStrokeColorRGB(0.8, 0.8, 0.8)
    c.setLineWidth(0.8)
    c.line(36, y - 68, W - 36, y - 68)

    # ── Item Table ──
    ty = y - 90
    c.setFillColorRGB(r, g, b)
    c.rect(36, ty, W - 72, 20, fill=1, stroke=0)
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(42, ty + 6, "Item / Description")
    c.drawString(310, ty + 6, "Qty")
    c.drawString(355, ty + 6, "Unit")
    c.drawString(400, ty + 6, "Unit Price (USD)")
    c.drawRightString(W - 42, ty + 6, "Extended Total (USD)")

    # Row 1
    ry = ty - 28
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(42, ry + 8, ITEM_NAME)
    c.setFont("Helvetica", 7.5)
    c.drawString(42, ry - 3, ITEM_SPEC)
    c.drawString(42, ry - 14, f"Model: {s['model']} | Manufacturer: {s['manufacturer']} | Origin: {s['country']}")

    qty   = s["qty"]
    price = s["unit_price"]
    total = price * qty

    c.setFont("Helvetica", 9)
    c.drawString(315, ry + 8, str(qty))
    c.drawString(355, ry + 8, "Units")
    c.drawString(400, ry + 8, f"${price:,.2f}")
    c.drawRightString(W - 42, ry + 8, f"${total:,.2f}")

    # Row line
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(36, ry - 24, W - 36, ry - 24)

    # Freight row
    freight_y = ry - 40
    c.setFont("Helvetica", 8.5)
    c.drawString(42, freight_y, "Freight / Shipping (DDP Houston, TX)")
    c.drawRightString(W - 42, freight_y, "Included")

    # Totals
    tot_y = freight_y - 28
    c.setFont("Helvetica-Bold", 10)
    c.drawString(350, tot_y, "TOTAL CONTRACT VALUE:")
    c.drawRightString(W - 42, tot_y, f"${total:,.2f}")

    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(350, tot_y - 14, "Applicable taxes not included.")

    # ── Commercial Terms Box ──
    box_y = tot_y - 55
    c.setFillColorRGB(0.96, 0.97, 0.99)
    c.rect(36, box_y - 140, W - 72, 150, fill=1, stroke=1)

    c.setFillColorRGB(r, g, b)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(48, box_y + 3, "COMMERCIAL TERMS & CONDITIONS:")

    terms = [
        ("Unit Price:",          f"${price:,.2f} USD per pump"),
        ("Extended Price:",      f"${total:,.2f} USD (12 units)"),
        ("Lead Time:",           f"{s['lead_time']} calendar days from PO confirmation"),
        ("Payment Terms:",       s["payment_terms"]),
        ("Incoterms:",           s["incoterms"]),
        ("Warranty:",            s["warranty"] + " from date of delivery"),
        ("Quote Validity:",      s["validity"] + " from date of issue"),
        ("Spare Parts:",         "Available — 2-year stocking agreement optional"),
        ("After-Sales Support:", "24/7 technical helpline | Site commissioning available"),
    ]

    term_y = box_y - 14
    for label, value in terms:
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(48, term_y, label)
        c.setFont("Helvetica", 8.5)
        c.drawString(200, term_y, value)
        term_y -= 15

    # ── AI Note / Compliance Note ──
    note_y = box_y - 158
    c.setFillColorRGB(0.98, 0.97, 0.88)
    c.rect(36, note_y - 28, W - 72, 38, fill=1, stroke=1)
    c.setFillColorRGB(0.6, 0.4, 0.0)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(48, note_y + 4, "NOTE:")
    c.setFont("Helvetica", 8)
    c.drawString(90, note_y + 4, s["note"])

    # ── Technical Compliance ──
    tc_y = note_y - 55
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(36, tc_y, "TECHNICAL COMPLIANCE STATEMENT:")
    c.setFont("Helvetica", 8)
    compliance_text = (
        f"{s['name']} confirms that the quoted equipment — {s['model']} — fully complies with the "
        f"technical specifications stated in {RFQ_REF}, including flow range (0–120 L/hr), "
        f"discharge pressure (≥7 bar), PVDF/PTFE wetted materials, 460V/3Ph/60Hz power supply, "
        f"4–20 mA control input, NEMA 4X enclosure, and ±2% accuracy. "
        f"Full datasheet, O&M manual, and technical drawings will be supplied upon PO award."
    )
    # Word-wrap manually
    words = compliance_text.split()
    line, lines = "", []
    for w in words:
        if c.stringWidth(line + w, "Helvetica", 8) < W - 100:
            line += w + " "
        else:
            lines.append(line.strip())
            line = w + " "
    lines.append(line.strip())
    for i, l in enumerate(lines):
        c.drawString(36, tc_y - 14 - i * 12, l)

    # ── Signature ──
    sig_y = 90
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.5)
    c.line(36, sig_y, 220, sig_y)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(36, sig_y - 14, "Authorized Signatory")
    c.setFont("Helvetica", 8)
    c.drawString(36, sig_y - 26, f"Sales & Commercial Team, {s['name']}")
    c.drawString(36, sig_y - 38, s["email"])

    # ── Footer ──
    c.setFillColorRGB(0.55, 0.55, 0.55)
    c.setFont("Helvetica-Oblique", 7)
    c.drawString(36, 35, f"Generated for demo: {s['label']} | {RFQ_REF} | Veolia Water Technologies")
    c.drawRightString(W - 36, 35, f"Page 1 of 1 — {QUOTE_DATE}")

    c.save()
    print(f"  [PDF] {file_path}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL GENERATOR
# ─────────────────────────────────────────────────────────────────────────────
def generate_xlsx(s, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Quotation"

    r8  = lambda h: int(h[0:2], 16)
    g8  = lambda h: int(h[2:4], 16)
    b8  = lambda h: int(h[4:6], 16)

    # Pick header color from supplier
    cr, cg, cb = s["color_rgb"]
    hex_color = "{:02X}{:02X}{:02X}".format(int(cr * 255), int(cg * 255), int(cb * 255))

    hdr_fill  = PatternFill(start_color=hex_color,  end_color=hex_color,  fill_type="solid")
    blue_fill = PatternFill(start_color="0078D4",   end_color="0078D4",   fill_type="solid")
    lt_fill   = PatternFill(start_color="F0F6FC",   end_color="F0F6FC",   fill_type="solid")
    note_fill = PatternFill(start_color="FFFBE6",   end_color="FFFBE6",   fill_type="solid")

    ft_title  = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ft_sec    = Font(name="Calibri", size=10, bold=True)
    ft_bold   = Font(name="Calibri", size=10, bold=True)
    ft_reg    = Font(name="Calibri", size=10)
    ft_hdr    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    ft_note   = Font(name="Calibri", size=9,  italic=True, color="7A4F00")

    thin = Side(style="thin", color="D0D5DD")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title Banner ──
    ws.merge_cells("A1:F2")
    ws["A1"].value     = f"SUPPLIER QUOTATION — {s['name'].upper()}"
    ws["A1"].font      = ft_title
    ws["A1"].fill      = hdr_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:F3")
    ws["A3"].value     = f"RFQ Reference: {RFQ_REF}  |  Source: {s['source']}  |  Quote Date: {QUOTE_DATE}"
    ws["A3"].font      = Font(name="Calibri", size=9, italic=True, color=hex_color)
    ws["A3"].alignment = Alignment(horizontal="center")

    # ── Vendor / Buyer ──
    ws["A5"].value = "VENDOR:"
    ws["A5"].font  = ft_sec
    for row, val in enumerate([
        s["name"], s["address"], s["email"], s["phone"]
    ], start=6):
        ws[f"A{row}"].value = val
        ws[f"A{row}"].font  = ft_reg

    ws["D5"].value = "PREPARED FOR:"
    ws["D5"].font  = ft_sec
    for row, val in enumerate([
        BUYER, DEPT, f"Delivery: {DELIVERY}", f"RFQ: {RFQ_REF}"
    ], start=6):
        ws[f"D{row}"].value = val
        ws[f"D{row}"].font  = ft_reg

    # ── Item Table ──
    hdrs = ["Item / Description", "Model", "Qty", "Unit", "Unit Price (USD)", "Extended Total (USD)"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=11, column=ci)
        cell.value     = h
        cell.font      = ft_hdr
        cell.fill      = blue_fill
        cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
        cell.border    = bdr
    ws.row_dimensions[11].height = 22

    qty   = s["qty"]
    price = s["unit_price"]
    total = price * qty

    row12 = [ITEM_NAME, s["model"], qty, "Units", price, total]
    for ci, v in enumerate(row12, 1):
        cell = ws.cell(row=12, column=ci)
        cell.value  = v
        cell.font   = ft_bold if ci in (1, 6) else ft_reg
        cell.border = bdr
        if ci in (5, 6):
            cell.number_format = '$#,##0.00'
            cell.alignment     = Alignment(horizontal="right")

    # Spec row
    ws.merge_cells("A13:F13")
    ws["A13"].value     = f"Spec: {ITEM_SPEC}"
    ws["A13"].font      = Font(name="Calibri", size=8.5, italic=True, color="555555")
    ws["A13"].alignment = Alignment(horizontal="left")

    # Freight
    ws["A14"].value = "Freight / Shipping"
    ws["A14"].font  = ft_reg
    ws["A14"].border = bdr
    ws["F14"].value = "Included"
    ws["F14"].font  = ft_reg
    ws["F14"].border = bdr
    ws["F14"].alignment = Alignment(horizontal="right")

    # Total
    ws.merge_cells("A16:E16")
    ws["A16"].value = "TOTAL CONTRACT VALUE (USD):"
    ws["A16"].font  = Font(name="Calibri", size=11, bold=True)
    ws["A16"].alignment = Alignment(horizontal="right")
    ws["F16"].value = total
    ws["F16"].font  = Font(name="Calibri", size=11, bold=True, color="0D5C20")
    ws["F16"].number_format = '$#,##0.00'
    ws["F16"].alignment = Alignment(horizontal="right")

    # ── Commercial Terms ──
    ws["A18"].value = "COMMERCIAL TERMS:"
    ws["A18"].font  = ft_sec

    terms = [
        ("Unit Price",         f"${price:,.2f} per pump"),
        ("Extended Price",     f"${total:,.2f} (12 units)"),
        ("Lead Time",          f"{s['lead_time']} calendar days from PO confirmation"),
        ("Payment Terms",      s["payment_terms"]),
        ("Incoterms",          s["incoterms"]),
        ("Warranty",           s["warranty"] + " from delivery date"),
        ("Quote Validity",     s["validity"]),
        ("Spare Parts",        "Available — 2-year stocking agreement optional"),
        ("After-Sales Support","24/7 technical helpline + site commissioning"),
        ("Country of Origin",  s["country"]),
    ]
    for ri, (lbl, val) in enumerate(terms, start=19):
        ws.cell(row=ri, column=1, value=lbl).font   = ft_bold
        ws.cell(row=ri, column=1).fill               = lt_fill
        ws.cell(row=ri, column=1).border             = bdr
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=6)
        cell = ws.cell(row=ri, column=2, value=val)
        cell.font   = ft_reg
        cell.border = bdr

    # ── Note ──
    note_row = 19 + len(terms) + 1
    ws.merge_cells(f"A{note_row}:F{note_row}")
    ws[f"A{note_row}"].value     = f"NOTE: {s['note']}"
    ws[f"A{note_row}"].font      = ft_note
    ws[f"A{note_row}"].fill      = note_fill
    ws[f"A{note_row}"].alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    # Column widths
    col_widths = [45, 25, 8, 8, 20, 22]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    wb.save(file_path)
    print(f"  [XLSX] {file_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n[START] Generating Veolia Demo Supplier Quotes -> {os.path.abspath(OUT_DIR)}\n")
    print(f"   RFQ: {RFQ_REF}")
    print(f"   Item: Chemical Dosing Pump Assemblies (12 units)\n")

    for s in SUPPLIERS:
        slug = s["name"].lower().replace(" ", "_")
        print(f"  >> {s['label']}: {s['name']}")

        # PDF
        pdf_path = os.path.join(OUT_DIR, f"quote_{slug}_veolia.pdf")
        generate_pdf(s, pdf_path)

        # Excel
        xlsx_path = os.path.join(OUT_DIR, f"quote_{slug}_veolia.xlsx")
        generate_xlsx(s, xlsx_path)
        print()

    print(f"\n[DONE]  {len(SUPPLIERS) * 2} files saved to:\n   {os.path.abspath(OUT_DIR)}\n")
    print("   Files:")
    for f in sorted(os.listdir(OUT_DIR)):
        size = os.path.getsize(os.path.join(OUT_DIR, f))
        print(f"      {f}  ({size:,} bytes)")
    print()
