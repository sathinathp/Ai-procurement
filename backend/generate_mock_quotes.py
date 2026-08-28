import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_supplier_quote_pdf(supplier_name, price, currency, lead_time, payment_terms, incoterms, item_name, file_path):
    c = canvas.Canvas(file_path, pagesize=letter)
    
    # 1. Header Banner
    c.setFillColorRGB(0.09, 0.17, 0.27) # Dark Slate Accent
    c.rect(0, 720, 612, 72, fill=1, stroke=0)
    
    c.setFillColorRGB(1.0, 1.0, 1.0)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(36, 750, supplier_name.upper())
    c.setFont("Helvetica", 9)
    c.drawString(36, 735, "OFFICIAL SUPPLIER QUOTATION & PROPOSAL")
    
    # Quotation Date and Number
    c.drawRightString(576, 750, f"Quote Ref: QT-2026-{supplier_name[:3].upper()}-98")
    c.drawRightString(576, 735, "Date: 2026-08-27")
    
    # 2. Bill To & Vendor Info
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(36, 680, "VENDOR DETAILS:")
    c.setFont("Helvetica", 9)
    c.drawString(36, 665, f"Company: {supplier_name}")
    c.drawString(36, 650, "Address: Industrial Zone Phase 2, Suite 410")
    c.drawString(36, 635, "Email: sales@" + supplier_name.lower().replace(" ", "") + ".com")
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(320, 680, "PREPARED FOR:")
    c.setFont("Helvetica", 9)
    c.drawString(320, 665, "Company: ProcureX Co. / Petabytz")
    c.drawString(320, 650, "Department: Global Procurement Operations")
    c.drawString(320, 635, "Project Ref: PRJ-2026-POMP-09")
    
    # Divider line
    c.setLineWidth(1)
    c.setStrokeColorRGB(0.8, 0.8, 0.8)
    c.line(36, 610, 576, 610)
    
    # 3. Itemized Quote Table Header
    y = 570
    c.setFillColorRGB(0.09, 0.57, 1.0) # ProcureX Blue accent
    c.rect(36, y, 540, 24, fill=1, stroke=0)
    
    c.setFillColorRGB(1.0, 1.0, 1.0)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(42, y + 8, "Item / Service Description")
    c.drawString(280, y + 8, "Qty")
    c.drawString(330, y + 8, "Unit")
    c.drawString(380, y + 8, "Unit Price")
    c.drawRightString(570, y + 8, f"Total ({currency})")
    
    # Item row details
    y -= 30
    c.setFillColorRGB(0.0, 0.0, 0.0)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(42, y + 8, item_name)
    c.setFont("Helvetica", 8.5)
    c.drawString(42, y - 4, f"Supplier Specification Match - Certified Model. Code: {supplier_name[:3].upper()}-99X")
    
    qty = 100.0 if "polymer" in item_name.lower() or "basf" in supplier_name.lower() or "sabic" in supplier_name.lower() or "khobar" in supplier_name.lower() or "borouge" in supplier_name.lower() else 5.0
    unit = "MT" if "polymer" in item_name.lower() or "basf" in supplier_name.lower() or "sabic" in supplier_name.lower() or "khobar" in supplier_name.lower() or "borouge" in supplier_name.lower() else "Pcs"
    
    total = price * qty
    
    c.setFont("Helvetica", 9)
    c.drawString(280, y + 8, f"{qty:.1f}")
    c.drawString(330, y + 8, unit)
    c.drawString(380, y + 8, f"{currency} {price:,.2f}")
    c.drawRightString(570, y + 8, f"{currency} {total:,.2f}")
    
    # Grid row line
    c.setLineWidth(0.5)
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(36, y - 12, 576, y - 12)
    
    # Subtotal and Total
    y -= 50
    c.setFont("Helvetica-Bold", 10)
    c.drawString(350, y, "Total Proposal Value:")
    c.drawRightString(570, y, f"{currency} {total:,.2f}")
    
    # 4. Commercial Terms & Conditions
    y -= 45
    c.setFillColorRGB(0.96, 0.97, 0.99)
    c.rect(36, y - 95, 540, 120, fill=1, stroke=1)
    
    c.setFillColorRGB(0.09, 0.17, 0.27)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(48, y + 10, "COMMERCIAL & LOGISTICS TERMS:")
    
    c.setFont("Helvetica", 9)
    terms = [
        ("Quoted Lead Time:", f"{lead_time} Calendar Days from Order Confirmation"),
        ("Payment Terms:", payment_terms),
        ("Incoterms (Delivery):", incoterms),
        ("Quotation Validity:", "30 Days from date of issue"),
        ("Quality Guarantee:", "12 Months standard manufacturer warranty")
    ]
    
    term_y = y - 10
    for title, desc in terms:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(48, term_y, title)
        c.setFont("Helvetica", 9)
        c.drawString(220, term_y, desc)
        term_y -= 18
        
    # Signatures
    y -= 150
    c.setFont("Helvetica-Bold", 9)
    c.drawString(36, y, "Authorized Signatory")
    c.line(36, y - 2, 180, y - 2)
    c.setFont("Helvetica", 8)
    c.drawString(36, y - 12, f"Sales Department, {supplier_name}")
    
    c.setFont("Helvetica-Oblique", 7.5)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawString(36, 40, f"Page 1 of 1 — Generated electronically via ProcureX simulation for {supplier_name}.")
    
    c.save()
    print(f"[OK] Generated supplier PDF quote: {file_path}")


def generate_supplier_quote_xlsx(supplier_name, price, currency, lead_time, payment_terms, incoterms, item_name, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Quotation"
    
    # Styling helpers
    navy_fill = PatternFill(start_color="172B43", end_color="172B43", fill_type="solid")
    blue_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F0F6FC", end_color="F0F6FC", fill_type="solid")
    gray_fill = PatternFill(start_color="F4F5F7", end_color="F4F5F7", fill_type="solid")
    
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="172B43")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_regular = Font(name="Calibri", size=10)
    font_table_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    border_thin_side = Side(style='thin', color='D0D5DD')
    border_double_bottom = Side(style='double', color='000000')
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    border_total = Border(top=border_thin_side, bottom=border_double_bottom)
    
    # 1. Title Banner
    ws.merge_cells("A1:E2")
    title_cell = ws["A1"]
    title_cell.value = f"QUOTATION PROPOSAL: {supplier_name.upper()}"
    title_cell.font = font_title
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Set Row Heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # 2. Metadata details
    ws["A4"] = "Vendor Details:"
    ws["A4"].font = font_section
    ws["A5"] = f"Company: {supplier_name}"
    ws["A5"].font = font_regular
    ws["A6"] = "Address: Industrial Zone Phase 2, Suite 410"
    ws["A6"].font = font_regular
    ws["A7"] = f"Contact: sales@{supplier_name.lower().replace(' ', '')}.com"
    ws["A7"].font = font_regular
    
    ws["D4"] = "Prepared For:"
    ws["D4"].font = font_section
    ws["D5"] = "Company: ProcureX Co. / Petabytz"
    ws["D5"].font = font_regular
    ws["D6"] = "Project Ref: PRJ-2026-POMP-09"
    ws["D6"].font = font_regular
    ws["D7"] = "Quote Date: 2026-08-27"
    ws["D7"].font = font_regular
    
    # 3. Item Table Header
    headers = ["Item Description", "Qty", "Unit", "Unit Price", f"Total ({currency})"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx)
        cell.value = header
        cell.font = font_table_hdr
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = border_thin
    
    ws.row_dimensions[9].height = 25
    
    # Item row details
    qty = 100.0 if "polymer" in item_name.lower() or "basf" in supplier_name.lower() or "sabic" in supplier_name.lower() or "khobar" in supplier_name.lower() or "borouge" in supplier_name.lower() else 5.0
    unit = "MT" if "polymer" in item_name.lower() or "basf" in supplier_name.lower() or "sabic" in supplier_name.lower() or "khobar" in supplier_name.lower() or "borouge" in supplier_name.lower() else "Pcs"
    total = price * qty
    
    ws.cell(row=10, column=1, value=item_name).font = font_bold
    ws.cell(row=10, column=1).border = border_thin
    
    ws.cell(row=10, column=2, value=qty).font = font_regular
    ws.cell(row=10, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=10, column=2).border = border_thin
    
    ws.cell(row=10, column=3, value=unit).font = font_regular
    ws.cell(row=10, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=10, column=3).border = border_thin
    
    ws.cell(row=10, column=4, value=price).font = font_regular
    ws.cell(row=10, column=4).number_format = f'$#,##0.00' if currency == 'USD' else f'[$€-2] #,##0.00'
    ws.cell(row=10, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=10, column=4).border = border_thin
    
    # Formula for Total
    total_cell = ws.cell(row=10, column=5, value="=B10*D10")
    total_cell.font = font_bold
    total_cell.number_format = f'$#,##0.00' if currency == 'USD' else f'[$€-2] #,##0.00'
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = border_thin
    
    # Totals Row
    ws.cell(row=12, column=4, value="Total Value:").font = font_bold
    ws.cell(row=12, column=4).alignment = Alignment(horizontal="right")
    
    final_total_cell = ws.cell(row=12, column=5, value="=E10")
    final_total_cell.font = Font(name="Calibri", size=11, bold=True)
    final_total_cell.number_format = f'$#,##0.00' if currency == 'USD' else f'[$€-2] #,##0.00'
    final_total_cell.alignment = Alignment(horizontal="right")
    final_total_cell.border = border_total
    
    # 4. Commercial Terms Section
    ws.cell(row=14, column=1, value="COMMERCIAL & SHIPPING TERMS:").font = font_section
    
    terms = [
        ("Quoted Lead Time", f"{lead_time} Calendar Days"),
        ("Payment Terms", payment_terms),
        ("Incoterms Delivery", incoterms),
        ("Warranty", "12 Months manufacturer warranty"),
        ("Quotation Validity", "30 Days from date of issue")
    ]
    
    start_row = 15
    for idx, (lbl, val) in enumerate(terms):
        ws.cell(row=start_row + idx, column=1, value=lbl).font = font_bold
        ws.cell(row=start_row + idx, column=1).fill = light_blue_fill
        ws.cell(row=start_row + idx, column=1).border = border_thin
        
        ws.cell(row=start_row + idx, column=2, value=val).font = font_regular
        ws.cell(row=start_row + idx, column=2).border = border_thin
        ws.merge_cells(start_row=start_row + idx, start_column=2, end_row=start_row + idx, end_column=4)
        for col in range(3, 5):
            ws.cell(row=start_row + idx, column=col).border = border_thin
            
    # Adjust column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(file_path)
    print(f"[OK] Generated supplier Excel quote: {file_path}")
